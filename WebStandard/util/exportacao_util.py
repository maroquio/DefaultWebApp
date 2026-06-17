"""
Utilitário de Exportação de Dados (CSV e Excel).

Fornece funções genéricas para exportar listas de dados para CSV e Excel,
adequadas para relatórios e downloads em qualquer módulo da aplicação.

Uso básico:
    from util.exportacao_util import exportar_csv, exportar_excel

    # No route:
    @router.get("/exportar-csv")
    async def exportar(request: Request, ...):
        dados = repo.obter_todos()
        campos = ["id", "nome", "email", "perfil"]
        headers = ["ID", "Nome", "E-mail", "Perfil"]
        return exportar_csv(dados, campos, headers, "usuarios.csv")

    # Excel (requer openpyxl: pip install openpyxl):
    return exportar_excel(dados, campos, headers, "usuarios.xlsx")

Dependências:
    - CSV: stdlib Python (nenhuma instalação necessária)
    - Excel: openpyxl (adicionar ao requirements.txt: openpyxl>=3.1.0)
"""

import csv
import io
import tempfile
from pathlib import Path
from typing import Any, Callable, Optional, Union
from datetime import datetime, date

from fastapi.responses import StreamingResponse, FileResponse

from util.logger_config import logger


# --------------------------------------------------------------------------- #
# CSV                                                                          #
# --------------------------------------------------------------------------- #

def exportar_csv(
    dados: list,
    campos: list[str],
    headers: Optional[list[str]] = None,
    nome_arquivo: str = "exportacao.csv",
    separador: str = ";",
    encoding: str = "utf-8-sig",
    converter: Optional[Callable[[Any, str], Any]] = None,
) -> StreamingResponse:
    """
    Exporta uma lista de dados para CSV como StreamingResponse.

    Eficiente em memória: gera o CSV linha por linha sem guardar tudo em memória.
    Usa utf-8-sig (BOM) por padrão para compatibilidade com Excel no Windows.

    Args:
        dados: Lista de objetos ou dicts com os dados
        campos: Lista de atributos/chaves a exportar (ex: ['id', 'nome', 'email'])
        headers: Cabeçalhos das colunas (ex: ['ID', 'Nome', 'E-mail'])
                 Se None, usa os nomes dos campos
        nome_arquivo: Nome do arquivo para download (ex: 'usuarios_2025.csv')
        separador: Separador de campos (';' padrão para compatibilidade com Excel BR)
        encoding: Encoding do arquivo ('utf-8-sig' inclui BOM para Excel Windows)
        converter: Função opcional para formatar valores: f(valor, nome_campo) -> str

    Returns:
        StreamingResponse com Content-Disposition para download

    Exemplo:
        from util.exportacao_util import exportar_csv

        @router.get("/usuarios/exportar-csv")
        @requer_autenticacao([Perfil.ADMIN.value])
        async def exportar_usuarios(request: Request, ...):
            usuarios = usuario_repo.obter_todos()
            return exportar_csv(
                dados=usuarios,
                campos=["id", "nome", "email", "perfil", "ativo"],
                headers=["ID", "Nome", "E-mail", "Perfil", "Ativo"],
                nome_arquivo="usuarios.csv",
            )
    """
    cabecalhos = headers or campos

    def _obter_valor(obj: Any, campo: str) -> Any:
        """Extrai valor de objeto ou dict."""
        if isinstance(obj, dict):
            return obj.get(campo, "")
        return getattr(obj, campo, "")

    def _formatar_valor(valor: Any, campo: str) -> str:
        """Formata valor para string amigável no CSV."""
        if converter:
            valor = converter(valor, campo)
        if valor is None:
            return ""
        if isinstance(valor, bool):
            return "Sim" if valor else "Não"
        if isinstance(valor, datetime):
            return valor.strftime("%d/%m/%Y %H:%M")
        if isinstance(valor, date):
            return valor.strftime("%d/%m/%Y")
        return str(valor)

    def _gerar_linhas():
        """Generator que produz linhas CSV uma por vez."""
        output = io.StringIO()
        writer = csv.writer(output, delimiter=separador, quoting=csv.QUOTE_MINIMAL)

        # BOM para Excel no Windows
        if encoding == "utf-8-sig":
            yield "\ufeff"

        # Cabeçalho
        writer.writerow(cabecalhos)
        yield output.getvalue()
        output.truncate(0)
        output.seek(0)

        # Dados
        for item in dados:
            linha = [_formatar_valor(_obter_valor(item, campo), campo) for campo in campos]
            writer.writerow(linha)
            yield output.getvalue()
            output.truncate(0)
            output.seek(0)

    logger.info(f"Exportando {len(dados)} registros para CSV: {nome_arquivo}")
    return StreamingResponse(
        content=_gerar_linhas(),
        media_type="text/csv; charset=utf-8-sig",
        headers={
            "Content-Disposition": f'attachment; filename="{nome_arquivo}"',
            "Cache-Control": "no-cache",
        },
    )


# --------------------------------------------------------------------------- #
# Excel                                                                        #
# --------------------------------------------------------------------------- #

def exportar_excel(
    dados: list,
    campos: list[str],
    headers: Optional[list[str]] = None,
    nome_arquivo: str = "exportacao.xlsx",
    titulo_aba: str = "Dados",
    converter: Optional[Callable[[Any, str], Any]] = None,
) -> FileResponse:
    """
    Exporta uma lista de dados para Excel (.xlsx) como FileResponse.

    Requer openpyxl instalado: pip install openpyxl>=3.1.0

    Funcionalidades:
    - Cabeçalhos em negrito com fundo colorido
    - Largura automática das colunas
    - Linhas alternadas com cor de fundo suave (zebra striping)
    - Formatação automática de datas, booleanos e números

    Args:
        dados: Lista de objetos ou dicts com os dados
        campos: Lista de atributos/chaves a exportar
        headers: Cabeçalhos das colunas. Se None, usa os nomes dos campos
        nome_arquivo: Nome do arquivo para download
        titulo_aba: Nome da aba da planilha
        converter: Função opcional para formatar valores: f(valor, nome_campo) -> Any

    Returns:
        FileResponse para download do arquivo .xlsx

    Raises:
        ImportError: Se openpyxl não estiver instalado

    Exemplo:
        from util.exportacao_util import exportar_excel

        @router.get("/usuarios/exportar-excel")
        @requer_autenticacao([Perfil.ADMIN.value])
        async def exportar_usuarios(request: Request, ...):
            usuarios = usuario_repo.obter_todos()
            return exportar_excel(
                dados=usuarios,
                campos=["id", "nome", "email", "perfil", "ativo"],
                headers=["ID", "Nome", "E-mail", "Perfil", "Ativo"],
                nome_arquivo="usuarios.xlsx",
                titulo_aba="Usuários",
            )
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "openpyxl é necessário para exportar Excel. "
            "Instale com: pip install openpyxl>=3.1.0"
        )

    cabecalhos = headers or campos

    def _obter_valor(obj: Any, campo: str) -> Any:
        if isinstance(obj, dict):
            return obj.get(campo, "")
        return getattr(obj, campo, "")

    def _formatar_valor(valor: Any, campo: str) -> Any:
        if converter:
            valor = converter(valor, campo)
        if valor is None:
            return ""
        if isinstance(valor, bool):
            return "Sim" if valor else "Não"
        return valor

    wb = Workbook()
    ws = wb.active
    ws.title = titulo_aba

    # Estilos do cabeçalho
    cor_cabecalho = "2563EB"  # Azul primário Bootstrap
    fonte_cabecalho = Font(bold=True, color="FFFFFF", size=11)
    preenchimento_cabecalho = PatternFill("solid", fgColor=cor_cabecalho)
    alinhamento_centro = Alignment(horizontal="center", vertical="center")

    # Escrita do cabeçalho
    for col_idx, header in enumerate(cabecalhos, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = fonte_cabecalho
        cell.fill = preenchimento_cabecalho
        cell.alignment = alinhamento_centro

    # Cor de fundo alternada para linhas
    cor_linha_par = "EFF6FF"  # Azul muito claro
    preenchimento_par = PatternFill("solid", fgColor=cor_linha_par)

    # Escrita dos dados
    for row_idx, item in enumerate(dados, start=2):
        for col_idx, campo in enumerate(campos, start=1):
            valor = _formatar_valor(_obter_valor(item, campo), campo)
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            if row_idx % 2 == 0:
                cell.fill = preenchimento_par

    # Ajustar largura das colunas automaticamente
    for col_idx, campo in enumerate(campos, start=1):
        letra = get_column_letter(col_idx)
        max_largura = len(str(cabecalhos[col_idx - 1]))
        for row_idx in range(2, len(dados) + 2):
            valor = ws.cell(row=row_idx, column=col_idx).value
            if valor:
                max_largura = max(max_largura, len(str(valor)))
        ws.column_dimensions[letra].width = min(max_largura + 4, 50)

    # Congelar painel do cabeçalho
    ws.freeze_panes = "A2"

    # Salvar em arquivo temporário e retornar
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    wb.save(tmp_path)
    logger.info(f"Exportando {len(dados)} registros para Excel: {nome_arquivo}")

    return FileResponse(
        path=tmp_path,
        filename=nome_arquivo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Cache-Control": "no-cache"},
        background=_excluir_arquivo_temp(tmp_path),
    )


class _excluir_arquivo_temp:
    """Background task para excluir o arquivo temporário após envio."""

    def __init__(self, path: str):
        self.path = path

    async def __call__(self):
        try:
            Path(self.path).unlink(missing_ok=True)
        except OSError:
            pass
